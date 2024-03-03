import streamlit as st
from jira import JIRA
import pandas as pd
from modules.config import JIRA_URL,TIMELINE_POWER_POINT_SLIDE,EXCEL_TIMELINE_ELEMENTS
from modules.jira_operations import get_company_managed_projects_df
import tempfile
from pptx import Presentation
from pptx.util import Inches, Pt
from pptx.enum.text import PP_ALIGN
from openpyxl import load_workbook
import shutil


# Get the excelfile and the slidedeck template
##########################################################################

# Path to excel template
template_path_excel = f'templates/ThinkCell/{EXCEL_TIMELINE_ELEMENTS}'
original_pptx_path = f'templates/ThinkCell/{TIMELINE_POWER_POINT_SLIDE}'
new_pptx_path = 'templates/ThinkCell/CUS_timeline.pptx'
# Copy the file
shutil.copyfile(original_pptx_path, new_pptx_path)
print(f"PowerPoint file copied to {new_pptx_path}")


# Create a copy using tempfile
##########################################################################


# Load the workbook and create a copy
template_workbook = load_workbook(template_path_excel)

# Save the workbook with the changes
template_workbook.save('templates/ThinkCell/CUS_timeline.xlsx')
copied_workbook = load_workbook('templates/ThinkCell/CUS_timeline.xlsx')
sheet = copied_workbook.active  # This selects the first sheet

# Read a cell's value
value = sheet['M2'].value
# Modify a cell's value
sheet['M2'] = '80' 

# Save files
copied_workbook.save('templates/ThinkCell/CUS_timeline.xlsx')



options = st.multiselect(
    'What are your favorite colors',
    ['Green', 'Yellow', 'Red', 'Blue'],
    ['Yellow', 'Red'])




