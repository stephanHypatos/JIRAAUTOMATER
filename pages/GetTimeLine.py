import streamlit as st
from jira import JIRA
import pandas as pd
from modules.config import JIRA_URL,TIMELINE_POWER_POINT_SLIDE,EXCEL_TIMELINE_ELEMENTS,EXCEL_TIMELINE_ELEMENTS_POC,TIMELINE_POWER_POINT_SLIDE_POC
from modules.jira_operations import get_jira_issue_type_project_key,save_jira_issue_type_project,get_project_keys,save_jira_project_key,get_all_jira_issues_of_project,get_due_date_by_summary,get_start_date_by_summary,save_jira_project_type
from modules.excel_operations import apply_named_style_and_fill_to_range
import tempfile
from pptx import Presentation
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle,PatternFill
from datetime import datetime
from modules.config import JIRA_URL
import zipfile
import io

st.set_page_config(page_title="Get Timelime Slide", page_icon=":calendar:")
st.title('Get Timeline Slide ')


# Initialize session state for JIRA API credentials if not already done
if 'api_username' not in st.session_state:
    st.session_state['api_username'] = ''
if 'api_password' not in st.session_state:
    st.session_state['api_password'] = ''
if 'jira_project_key' not in st.session_state:
    st.session_state['jira_project_key'] = ''
if 'jira_project_type' not in st.session_state:
    st.session_state['jira_project_type'] = ''
if 'jira_issue_type_project' not in st.session_state:
    st.session_state['jira_issue_type_project'] = ''
if 'issue_data' not in st.session_state:
    st.session_state['issue_data'] = ''
if 'zip_buffer' not in st.session_state:
    st.session_state['zip_buffer'] = None


jira = JIRA(JIRA_URL, basic_auth=(st.session_state['api_username'], st.session_state['api_password']))
# Get Project Keys from Jira
project_keys = get_project_keys(JIRA_URL, st.session_state['api_username'], st.session_state['api_password'])

with st.expander('Read the Documentation on how to work with the Timeline Files'):
    st.markdown(
        """ 
        [Documentation](https://hypatos.atlassian.net/wiki/spaces/PD/pages/1300037695/JIRA+to+Powerpoint+Timeline+with+HYPA+PMO)
        """
        )
# Select Project Key
project = st.selectbox("Select Jira Board", project_keys, index=0)
with st.expander('Expand for more info on how to find the Jira Board Key'):
    st.write('You can find the key of your Jira board by clicking on DisplayBoards on the left pane.')
save_jira_project_key(project)

# Get all Issues of Type "Project" in a given Jira Board

if st.session_state['jira_project_key']:
    jira_projects = get_jira_issue_type_project_key(JIRA_URL, st.session_state['api_username'], st.session_state['api_password'])
    # Select Project Parent (most likely an Issue Type Account)
    project = st.selectbox("Select a Project from the given Jira Board", jira_projects, index=0)
    save_jira_issue_type_project(project)
    
    project_type = st.selectbox("Select the Type of your Project", ["POC", "PILOT"], index=0)
    save_jira_project_type(project_type)



# Create a copy of temp excel file
def copy_excel(excel_template_name):
    
    # Path to excel template
    template_path_excel = f'templates/ThinkCell/{excel_template_name}'
    # Load the workbook and create a copy
    template_workbook = load_workbook(template_path_excel)
    sheet = template_workbook.active  # This selects the first sheet
    
    # Update the Cells with the dates from the Jira Project 
    
    # MAP Dates of EPICs to excel file
    # User Training
    issue_summary='User Training'
    sheet['B44']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    sheet['C44']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)
    
    # Project Management  
    issue_summary='Project Management'
    sheet['B45']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    
    # Assessment 
    issue_summary='Assessment'
    sheet['B46']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    sheet['C46']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)
    
    # Design 
    issue_summary='Design'
    sheet['B47']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    sheet['C47']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)
    
    # Machine Learning 
    issue_summary='Machine Learning'
    sheet['B48']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    sheet['C48']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)

    # Application Implementation
    issue_summary='Application Implementation'
    sheet['B49']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    sheet['C49']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)

    if st.session_state['jira_project_type'] =='PILOT':
        # Integration Implementation
        issue_summary='Integration Implementation'
        sheet['B50']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
        sheet['C50']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)

    # Testing
    issue_summary='Testing'
    sheet['B51']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    sheet['C51']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)

    # Delivery
    issue_summary='Delivery'
    sheet['B52']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    sheet['C52']=get_due_date_by_summary(st.session_state['issue_data'], issue_summary)

    # MAP Dates of Workshops
    
    # Perform Assessment Workshop
    issue_summary='Perform Assessment Workshop'
    sheet['C34']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)

    # Perform Functional Workshop
    issue_summary='Perform Functional Workshop'
    sheet['C35']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    
    if st.session_state['jira_project_type'] =='PILOT':
        # Perform Technical Workshop
        issue_summary='Perform Technical Workshop'
        sheet['C36']=get_start_date_by_summary(st.session_state['issue_data'], issue_summary)
    
    # Apply  formatting to the cells
    date_style = NamedStyle(name="date_style", number_format="DD.MM.YY")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    date_style = NamedStyle(name="date_style", number_format="DD.MM.YY")
    apply_named_style_and_fill_to_range(sheet, "C34:C36", date_style, yellow_fill)
    apply_named_style_and_fill_to_range(sheet, "B44:B52", date_style, yellow_fill)
    
    # Save the presentation to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpfile:
        template_workbook.save(tmpfile.name)
        return tmpfile.name

# Create a copy of temp powerpoint file
def copy_powerpoint(slide_template_name):
    
    # Path to template and output PowerPoint files (Weekly Report)
    original_pptx_path = f'templates/ThinkCell/{slide_template_name}'
    # Load PowerPoint template
    presentation = Presentation(original_pptx_path)


    # Save the presentation to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pptx") as tmpfile:
        presentation.save(tmpfile.name)
        return tmpfile.name

# Download the powerpoint and excel files zipped
def download_files_as_zip(slide_template_name,excel_template_name):
    # Create the PowerPoint and Excel files
    presentation_path = copy_powerpoint(slide_template_name)
    excel_path = copy_excel(excel_template_name)

    # Create a ZIP file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        zip_file.write(presentation_path, arcname="timeline_ppt_copy.pptx")
        zip_file.write(excel_path, arcname="timeline_xlsx_copy.xlsx")

    # Return to the beginning of the stream
    zip_buffer.seek(0)
    
    return zip_buffer

def clear_zip_buffer():
    st.session_state['zip_buffer'] = None
    st.session_state['jira_project_type'] = None
    st.session_state['issue_data'] = None



if st.button("Get Updated Timeline Slide"):
    if st.session_state['api_username'] and st.session_state['api_password']:
        jira = JIRA(JIRA_URL, basic_auth=(st.session_state['api_username'], st.session_state['api_password']))
    else:
        st.warning("Please provide Jira credentials.")

    if not st.session_state['jira_project_key']:
        st.warning("Please select Jira Board Key.")
    if not st.session_state['jira_issue_type_project']:
        st.warning("Please select Jira Project Key.")
    
    else:
        try:
            with st.spinner('In progress... dont worry, this can take up to 2 mins.'):
                st.session_state['issue_data'] = get_all_jira_issues_of_project(jira, st.session_state['jira_issue_type_project'])
                if st.session_state['jira_project_type'] == 'PILOT':
                    # Generate Files for Pilot
                    st.session_state['zip_buffer']=download_files_as_zip(TIMELINE_POWER_POINT_SLIDE, EXCEL_TIMELINE_ELEMENTS)
                   
                else: 
                    # Generate Files for PoC 
                    st.session_state['zip_buffer']=download_files_as_zip(TIMELINE_POWER_POINT_SLIDE_POC, EXCEL_TIMELINE_ELEMENTS_POC)
            
            st.success("Files are ready for download.")
        except Exception as e:
            st.warning(f"Error Msg: {e}")

# Display download button if zip_buffer is available
if st.session_state['zip_buffer']:
    st.download_button(
        label="Download Timeline Files",
        data=st.session_state['zip_buffer'],
        file_name="timeline_files.zip",
        mime="application/zip",
        on_click=clear_zip_buffer
    ) 